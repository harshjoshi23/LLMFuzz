"""Document ingestion for production CLI.

Supports offline ingestion from:
- local folders (markdown/txt/pdf)
- exported Confluence pages (HTML or TXT)
- exported Jira pages (HTML or TXT)

Design: produce a unified corpus folder per project:
- <index_dir>/corpus/*.txt
and build RAG index from that folder.

NOTE: This is intentionally offline-first for thesis demos.
"""

from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import List


@dataclass
class IngestResult:
    files_copied: int
    corpus_dir: str


SUPPORTED_EXTS = {".md", ".txt", ".pdf", ".html", ".htm", ".docx", ".xlsx", ".rst"}


def _xlsx_to_text(src: Path, dst: Path) -> bool:
    """Render an .xlsx file as a flat .txt corpus chunk.

    Returns True on success, False if openpyxl is unavailable or read fails.
    The output filename uses the original stem plus ``.xlsx.txt`` so the
    downstream chunker treats it as text.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return False
    try:
        wb = load_workbook(filename=str(src), data_only=True, read_only=True)
    except Exception:
        return False

    lines: List[str] = [f"# Source: {src.name}"]
    for ws in wb.worksheets:
        lines.append(f"\n## Sheet: {ws.title}")
        header: List[str] = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            cells = ["" if v is None else str(v).strip() for v in row]
            if not any(cells):
                continue
            if row_idx == 0:
                header = cells
                lines.append(" | ".join(cells))
                continue
            if header and len(header) == len(cells):
                pairs = [f"{h}={v}" for h, v in zip(header, cells) if v and h]
                lines.append("; ".join(pairs))
            else:
                lines.append(" | ".join(cells))
    out_text = "\n".join(lines)
    dst.write_text(out_text, encoding="utf-8")
    return True



class DocsIngestor:
    def __init__(self, project_root: str):
        self.project_root = Path(project_root)
        # Repo root = directory containing this src/ package's parent.
        # Used as a secondary search root for doc paths that don't exist
        # under project_root (e.g. shared data/datasheets folders).
        self.repo_root = Path(__file__).resolve().parents[1]

    def _resolve_source(self, src: str) -> Path:
        p = Path(src)
        if p.is_absolute():
            return p
        # Try project root first.
        cand_proj = (self.project_root / p).resolve()
        if cand_proj.exists():
            return cand_proj
        # Fall back to repo root, but ONLY for paths that point at a specific
        # subfolder/file (never for "." or ""). This avoids accidentally
        # ingesting the entire repository when target_path is missing.
        if str(p) not in (".", "", "./"):
            cand_repo = (self.repo_root / p).resolve()
            if cand_repo.exists():
                return cand_repo
        return cand_proj

    def ingest(self, sources: List[str], corpus_dir: str) -> IngestResult:
        out = Path(corpus_dir)
        out.mkdir(parents=True, exist_ok=True)

        copied = 0
        seen_names: set[str] = set()

        def _unique_name(base: str) -> str:
            # If name collides (common for nested CMakeLists.txt/README.md),
            # disambiguate with a numeric suffix instead of silently overwriting.
            if base not in seen_names:
                seen_names.add(base)
                return base
            stem = Path(base).stem
            suf = Path(base).suffix
            i = 1
            while True:
                cand = f"{stem}__{i}{suf}"
                if cand not in seen_names:
                    seen_names.add(cand)
                    return cand
                i += 1

        for src in sources:
            p = self._resolve_source(src)

            if not p.exists():
                continue

            if p.is_file():
                if p.suffix.lower() == ".xlsx":
                    if _xlsx_to_text(p, out / _unique_name(p.stem + ".xlsx.txt")):
                        copied += 1
                    continue
                if p.suffix.lower() in SUPPORTED_EXTS:
                    shutil.copy2(p, out / _unique_name(p.name))
                    copied += 1
                continue

            # directory
            for f in p.rglob("*"):
                if not f.is_file():
                    continue
                ext = f.suffix.lower()
                if ext == ".xlsx":
                    if _xlsx_to_text(f, out / _unique_name(f.stem + ".xlsx.txt")):
                        copied += 1
                    continue
                if ext in SUPPORTED_EXTS:
                    # Build a path-aware slug so nested CMakeLists.txt /
                    # README.md files do not collide.
                    try:
                        rel_parts = f.relative_to(p).parts
                    except ValueError:
                        rel_parts = (f.name,)
                    if len(rel_parts) <= 1:
                        slug = f.name
                    else:
                        slug = "_".join(rel_parts[:-1]) + "_" + f.name
                    shutil.copy2(f, out / _unique_name(slug))
                    copied += 1

        return IngestResult(files_copied=copied, corpus_dir=str(out))
